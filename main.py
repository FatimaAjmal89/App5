import pandas as pd
import glob
from fpdf import FPDF
from pathlib import Path

from openpyxl.styles.builtins import total

#pandas need openpyxl to read excel files
#xlsx is a used for reading excel sheets

filepaths = glob.glob("invoices/*.xlsx")
for filepath in filepaths:

    pdf = FPDF(orientation="P",unit="mm",format="A4")
    pdf.add_page()

    filename = Path(filepath).stem
    invoices_nr,date = filename.split("-")

    pdf.set_font(family="Times",size=16,style="B")
    pdf.cell(w=50 ,h=8 ,txt= f"Invoice nr.{invoices_nr}",ln=1 )

    pdf.set_font(family="Times", size=16, style="B")
    pdf.cell(w=50, h=8, txt=f"Date: {date}",ln=1 )

    df = pd.read_excel(filepath, sheet_name="Sheet 1")
    column = df.columns
    column = [item.replace("_"," ").title() for item in column]

    pdf.set_font(family="Times", size=8, style="B")
    pdf.set_text_color(80, 80, 80)
    pdf.cell(w=30, h=8, txt=column[0], border=1)
    pdf.cell(w=70, h=8, txt=column[1], border=1)
    pdf.cell(w=30, h=8, txt=column[2], border=1)
    pdf.cell(w=30, h=8, txt=column[3], border=1)
    pdf.cell(w=30, h=8, txt=column[4], ln=1, border=1)


    for index,row in df.iterrows():
        pdf.set_font(family="Times", size=8)
        pdf.set_text_color(80,80,80)
        pdf.cell(w=30 ,h=8 ,txt=str(row["product_id"]),border=1)
        pdf.cell(w=70, h=8, txt=str(row["product_name"]),border=1)
        pdf.cell(w=30, h=8, txt=str(row["amount_purchased"]),border=1)
        pdf.cell(w=30, h=8, txt=str(row["price_per_unit"]),border=1)
        pdf.cell(w=30, h=8, txt=str(row["total_price"]),ln=1,border=1)

    total_sum = df["total_price"].sum()
    pdf.set_font(family="Times", size=8)
    pdf.set_text_color(80, 80, 80)
    pdf.cell(w=30, h=8, txt="", border=1)
    pdf.cell(w=70, h=8, txt="", border=1)
    pdf.cell(w=30, h=8, txt="", border=1)
    pdf.cell(w=30, h=8, txt="", border=1)
    pdf.cell(w=30, h=8, txt=str(total_sum), ln=1, border=1)

    pdf.set_font(family="Times", size=8 ,style="B")
    pdf.set_text_color(80, 80, 80)
    pdf.cell(w=30, h=8, txt=f"The total sun is {total_sum}",ln=1)

    pdf.set_font(family="Times",size=14 ,style="B")
    pdf.set_text_color(0, 0, 0)
    pdf.cell(w=25, h=8, txt="PythonHow")
    pdf.image("pythonhow.png", w=10)





    pdf.output(f"PDFS/{filename}.pdf")
